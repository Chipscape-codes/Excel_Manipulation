import openpyxl as op
from openpyxl.styles import Font, Alignment
import excel_as as ea


a = op.load_workbook('test.xlsx')
F = input("Enter the title of the file: ")
b = int(input("Enter the number of rows(||): "))
c = int(input("Enter the number of columns(==): "))
a.active.title = F
sheet = a.active
sheet.cell(row=1, column=1).value = "%s"%(F)
sheet.cell(row=1, column=1).font = Font(bold=True)
sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

d = []
for i in range(c):
  e = input("Enter the column name: ")
  sheet.cell(row=1, column=i+2).value = e
  sheet.cell(row=1, column=i+2).font = Font(bold=True)
  sheet.cell(row=1, column=i+2).alignment = Alignment(horizontal='center')
  d.append(e)

for i in range(b):
  sheet.cell(row=i+2, column=1).value = i+1
  sheet.cell(row=i+2, column=1).alignment = Alignment(horizontal='center')
  for j in range(c):
    sheet.cell(row=i+2, column=j+2).value = input(f"Enter the value for row {i+1}, column {d[j]}: ")

J=input("Do you want to perform a addition of subtraction? (y/n): ")
if(J == 'y'):
  ea.arithmatic(J,b)
else:
  pass  

a.save(f'{F}.xlsx')
