def arithmetic(J,b):
    if J == 'y':
        import openpyxl as op
        d=input("do you want to perform addition or subtraction? (a/s): ")
        if d == 'a':
            a = op.load_workbook('test.xlsx')
            sheet = a.active
            E=input("do you want to add column or row? (c/r):")
            if E == 'c':
                b = int(input("Enter the number of rows(||): "))
                c = int(input("Enter the number of columns(==): "))
                for i in range(2, b+1):
                    sheet.cell(row=i, column=c+1).value = f'=SUM(B{i}:B{b})'
            elif E == 'r':
                b = int(input("Enter the number of rows(||): "))
                c = int(input("Enter the number of columns(==): "))
                for i in range(2, c+1):
                    sheet.cell(row=b+1, column=i).value = f'=SUM({chr(65+i)}2:{chr(65+i)}{c})'
        elif d == 's':
            a = op.load_workbook('test.xlsx')
            sheet = a.active
            d=input("do you want to subtract column or row? (c/r): ")
            if d == 'c':
                b = int(input("Enter the number of rows(||): "))
                c = int(input("Enter the number of columns(==): "))
                for i in range(2, b+1):
                    sheet.cell(row=i, column=c+1).value = f'=B{i}-B{b}'
            elif d == 'r':
                c = int(input("Enter the position of column(==): "))
                for i in range(2, c+1):
                    total=sheet.cell(row=b+1, column=i).value = f'=b{i}-b{c}'   
                sheet.cell(row=b+1, column=i).value = total
    else:
        pass